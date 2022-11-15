from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from django.utils import timezone
from university.celery import app
from studies.models import Faculty, Student, EducationGroups
from django.core.mail import EmailMessage


@app.task(name="CreateStudents")
def send_email_students_created(subject, from_email, to, body):
    msg = EmailMessage(
        subject=subject,
        from_email=from_email,
        to=to,
        body=body,
    )
    msg.send(fail_silently=True)

    return None


@app.task(name="Report")
def report():
    qs = (
        Faculty.objects.prefetch_related("subjects_of_study")
        .select_related("curator")
        .all()
    )

    wb = Workbook()
    ws = wb.active

# оформление таблицы

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    ft = Font(color="FF0000", size=14)
    br_top = Border(top=double, left=double, right=double, bottom=double)
    br = Border(top=thin, left=double, right=double, bottom=thin)

    for column in range(1, 4):
        cell = ws.cell(1, column)
        cell.font = ft
        cell.border = br_top

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60

# заполнение таблицы

    ws.cell(row=1, column=1, value="Направление")
    ws.cell(row=1, column=2, value="Куратор")
    ws.cell(row=1, column=3, value="Предметы")

    rn = 2
    for faculty in qs:
        ws.cell(row=rn, column=1, value=faculty.name)
        ws.cell(row=rn + 0, column=2, value=faculty.curator.last_name)
        ws.cell(row=rn + 1, column=2, value=faculty.curator.first_name)
        ws.cell(row=rn + 2, column=2, value=faculty.curator.patronymic_name)
        ws.cell(row=rn + 3, column=2, value=faculty.curator.passport)
        k = 0
        for sbjct in faculty.subjects_of_study.all():
            ws.cell(row=rn + k, column=3, value=sbjct.name)
            k += 1
        rn += max(3, k) + 1

# оформление таблицы

    for row in range(2, rn-1):
        for column in range(1, 4):
            cell = ws.cell(row, column)
            cell.border = br

############################################################################

    filename = "Report{}.xls".format(timezone.now())
    wb.save(filename)


@app.task(name="Report_groups")
def report_2():
    groups = EducationGroups.objects.all()

    wb = Workbook()
    ws = wb.active

# оформление таблицы

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30


    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    ft = Font(color="FF0000", size=14)
    br_top = Border(top=double, left=double, right=double, bottom=double)
    br = Border(top=thin, left=double, right=double, bottom=thin)

    for column in range(1, 6):
        cell = ws.cell(1, column)
        cell.font = ft
        cell.border = br_top

# заполнение таблицы

    ws.cell(row=1, column=1, value="Группа")
    ws.cell(row=1, column=2, value="Список студентов")
    ws.cell(row=1, column=3, value="Количество мужчины")
    ws.cell(row=1, column=4, value="Количество женщин")
    ws.cell(row=1, column=5, value="Количество свободных мест")

    rn = 2
    for item in groups:
        qs = Student.objects.values(
            "student_user__last_name",
            "student_user__first_name",
            "student_user__patronymic_name",
            "student_user__passport",
            "student_user__gender",
            "group__name",
        ).filter(group__name=item.name)

        k = 0
        ws.cell(row=rn, column=1, value=item.name)
        for i in qs:
            ws.cell(row=rn + k, column=2, value=f'{i["student_user__last_name"]} {i["student_user__passport"]}')
            k += 1
        m = qs.filter(student_user__gender="M").count()
        f = qs.filter(student_user__gender="F").count()
        ws.cell(row=rn, column=3, value=m)
        ws.cell(row=rn, column=4, value=f)
        ws.cell(row=rn, column=5, value=20-m-f)

        rn += max(1, k) + 1

# оформление таблицы

    for row in range(2, rn-1):
        for column in range(1, 6):
            cell = ws.cell(row, column)
            cell.border = br

#######################################################

    filename = "Report_groups{}.xls".format(timezone.now())
    wb.save(filename)
